import openpyxl as excel
from openpyxl import load_workbook


def write_excel(data_list, file_name):
    write_excel_title(data_list, file_name, [""])


def write_excel_title(data_list, file_name, title):
    """
       data_list 数据列表
       file_name 文件名称
       """
    wb = excel.Workbook()
    ws = wb.active
    ws.title = '数据处理'
    # names = ["column1", "column2", "column3", "column4", "column5", "column6", "column7"]
    ws.append(title)
    for ele in data_list:
        ws.append(ele)
    wb.save(file_name)


def read_file_tmp(file):
    wk_book = load_workbook(filename=file)
    sheet_name = "Sheet1"
    sheet = wk_book[sheet_name]
    # 获取指定范围内的单元格区域
    for row in sheet['A1:E8']:
        for cell in row:
            print(cell.value)

    # 获取所有列的值
    for col in sheet.iter_cols(values_only=True):
        for cell_value in col:
            print(cell_value)

    # 获取所有行的值
    for row in sheet.iter_rows(values_only=True):
        for cell_value in row:
            print(cell_value)

    # 使用 iter_cols() 方法获取指定范围内的一系列列，并遍历每列中的每个单元格
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=1):
        for cell in col:
            print(cell.value)

    for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1):
        for cell in row:
            print(cell.value)


app_names = [

    "234",
    "345"
]


def read_excel(file):
    # r
    wk_book = load_workbook(filename=file)
    sheet_name = "数据导出"
    sheet = wk_book[sheet_name]
    data_list = []
    app_set = set()
    # 获取所有行的值
    for row in sheet.iter_rows(values_only=True):
        app = row[6]
        if app not in app_names:
            continue
        # app_set.add(app)
        tmp = [row[6], row[7], row[8], row[9], row[14]]
        data_list.append(tmp)
        # for cell_value in row:
        #     print(cell_value)
    write_excel(data_list, "./2024032911.xlsx")


if __name__ == '__main__':
    print("handle excel ! ")
    read_excel("./2024032910.xlsx")
